import os
import argparse
from dpsubsystem import DPSubsystem
from dpslave import DPSlave
from slot import Slot
import re
import logging
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

is_Subsystem = r'^DPSUBSYSTEM{1}(?!.*DPADDRESS)(?!.*SLOT).*$'
is_Slave = r'^DPSUBSYSTEM{1}(.*DPADDRESS)(?!.*SLOT).*$'
is_Slot = r'^(?!MASTER)DPSUBSYSTEM{1}(.*DPADDRESS)(.*SLOT).*$'
is_Type_Address = r'^LOCAL(.*ADDRESS).*$'
is_Address = r'^(?!LOCAL)(.*ADDRESS).*$'
is_End = r'^(END)'

def create_hw_tree(text):
    logger.info("Starting creation of hardware tree from text with {} lines.".format(len(text)))
    hw_tree = {}

    #last_object_create = "none"

    for i in range(len(text)):
        line = text[i].rstrip()
        
        if (re.match(is_Subsystem, line) is not None):
            init = i
            while (re.match(is_End, text[i]) is None):
                i += 1
            end = i
            new_subsystem = DPSubsystem(text[init:end]) 
            hw_tree[new_subsystem.get_id] = new_subsystem
        elif (re.match(is_Slave, line) is not None):
            init = i
            while (re.match(is_End, text[i]) is None):
                i += 1
            end = i
            new_slave = DPSlave(text[init:end])
            hw_tree[new_slave.get_subsystem_id].set_new_slave(new_slave)
        elif (re.match(is_Slot, line) is not None):
            init = i
            while (re.match(is_End, text[i]) is None):
                i += 1
            end = i
            new_slot = Slot(text[init:end])
            hw_tree[new_slot.get_subsystem_id].get_slave(new_slot.get_slave_id).set_new_slot(new_slot)

    return hw_tree

def main():
    parser = argparse.ArgumentParser(description="Analize a Simatic hardware configuration file.")
    parser.add_argument('filename', type=str, help="Path to the file (*.cfg)")
    args = parser.parse_args()

    logger.info("Opening file: {}.".format(args.filename))
    file_to_read = open(args.filename,'r')
    try:
        lines = file_to_read.readlines()
    except UnicodeDecodeError as ude:
        lines = ""
        print(f'has occurred an error: {ude}')
    
    tree = create_hw_tree(lines)
    file_to_read.close()

    logger.info("Printing hardware description with {} subsystems.".format(len(tree)))
    
    data = list()
    for sys in tree:
        data_sys = tree[sys].description
        for addr in data_sys:
            data.append(addr)

    #get directory from args.filename
    directory = os.path.dirname(args.filename)
    #create a new file with the same name but with .xlsx extension
    new_file = os.path.join(directory, os.path.splitext(os.path.basename(args.filename))[0] + ".xlsx")
    #Fill openpyxl file with the data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simatic"
    #Find dictionary most long in data
    max_len = 0
    max_index = 0
    for i in data:
        if len(i) > max_len:
            max_len = len(i)
            max_index = i
    #Get columns titles from the most long dictionary
    columns = list(max_index.keys())
    #Add columns titles to the first row
    for i in range(len(columns)):
        ws.cell(row=1, column=i+1, value=columns[i])
    #Fill the rest of the file with the data using the columns titles as keys. if key is not in dictionary, add empty string
    for i in range(len(data)):
        for j in range(len(columns)):
            if columns[j] in data[i]:
                #Remove line breaks from the string
                ws.cell(row=i+2, column=j+1, value=data[i][columns[j]].replace('\n', ' '))
            else:
                ws.cell(row=i+2, column=j+1, value="")    
        
    #Save the file
    wb.save(new_file)
    logger.info("File saved as {}.".format(new_file))

    
    



    # df = pd.DataFrame(data=data)
    # df = df[['red', 'dir_profibus','nombre_equipo','ref_equipo','comentario','slot_modulo','ref_modulo','nombre_modulo','dir', 'tag','descripcion','tipo','rango']]

    # output_file_name =  os.path.dirname(os.path.abspath(args.filename)) + "/HW_results.csv"
    # logger.info("Writing hardware description in {} file.".format(output_file_name))
    # df.to_csv(output_file_name, index=False)

if __name__ == "__main__":
    main()